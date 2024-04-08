# -*- coding: utf-8 -*-
"""
Created on Thu Sep  1 15:45:32 2022

@author: yangxy
"""


import pandas as pd
import numpy as np
import copy
import openpyxl as op
import os
import xlwings as xw

app=xw.App(visible=False,add_book=False)
app.display_alerts=False
app.screen_updating=False


#一、先读取fundhold

#os.chdir(r"D:\desktop\index_research\datamix\data")
dirpath=r"D:\desktop\mydatabase\fundhold"
dirlist=os.listdir(dirpath)

'''
hang=len(wb.sheets[0].range('A1').current_region.rows)
lie=len(wb.sheets[0].range('A1').current_region.columns)
df=wb.sheets[0].range((1,1),(hang,lie)).options(pd.DataFrame,index=False).value
'''

greatlis=dict()
ok=dict()
for filename in dirlist:
    print(filename)
    #break
    if os.path.isdir(os.path.join(dirpath,filename)):
        continue
    elif  filename=="ok.xlsx":
        wb=app.books.open(os.path.join(dirpath,filename))
        wb.sheets[0].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[0].range('A1').options(pd.DataFrame,index=False,expand="table").value
        ok["重仓股代码"]=df
        wb.close()
        continue
    wb=app.books.open(os.path.join(dirpath,filename))
    #wb[wb.sheetnames[0]].title
    greatdf=dict()
    for she in range(len(wb.sheets)):
        wb.sheets[she].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[she].range('A1').options(pd.DataFrame,index=False,expand="table").value
        greatdf[wb.sheets[she].name]=df
    greatlis[filename]=greatdf
    wb.close()


tmp1=set()
for i in greatlis.keys():
    tmp1=set(greatlis[i].keys()) | tmp1
shtlis=copy.deepcopy(tmp1)
del tmp1
holddic=dict()
for sht in shtlis:
    #break
    df=pd.DataFrame(columns=["date"])
    for sce in greatlis.values():
        #break
        if sht in sce.keys():
            tmp1=copy.deepcopy(sce[sht])
            tmp1.index=pd.to_datetime(tmp1['date'])
            tmp1=tmp1.resample('3M',axis=0,closed="right",label="right").last()
            tmp1['date']=tmp1.index
            tmp1.index.name='index'
            df=pd.merge(df,tmp1,how="outer",on="date")
            tmp1=df.copy()
            i=sht
            tmp2=pd.Series(tmp1.columns).apply(lambda x:True if x[-2:-1]=="_" else False)
            if len(tmp2[tmp2])>0:
                tmp3=tmp1.loc[:,tmp1.columns[tmp2]]
                tmp4=pd.Series(tmp3.columns).apply(lambda x:x[:-2])
                tmp4=set(tmp4)
                tmp7=pd.DataFrame(index=tmp1.index)
                for j in tmp4:
                    #break
                    tmp5=tmp3.loc[:,j+"_x"]
                    tmp6=tmp3.loc[:,j+"_y"]
                    tmp5=tmp5.replace(0,np.nan)
                    tmp6=tmp6.replace(0,np.nan)
                    for k in range(len(tmp5)):
                        #break
                        if (tmp5[k]==tmp5[k]) or (tmp5[k]==""):
                            pass
                        else:
                            tmp5[k]=tmp6[k]
                    tmp5=pd.DataFrame(tmp5,index=tmp1.index)
                    tmp5.columns=[j]
                    tmp7=pd.concat([tmp7,tmp5],axis=1)
                tmp1=tmp1.loc[:,tmp1.columns[tmp2.apply(lambda x:not(x))]]
                del tmp2,tmp3,tmp4,tmp5,tmp6
                tmp7=pd.concat([tmp1['date'],tmp7],axis=1)
                tmp1=pd.concat([tmp7,tmp1.iloc[:,1:]],axis=1)
                df=tmp1.copy()
        else:
            pass
    df=df.sort_values(by='date',axis=0,ascending=True,inplace=False,na_position='last')
    df=df.reset_index(drop=True)
    holddic[sht]=df
del sht,sce,tmp1
for k in holddic.keys():
    #break
    tmp1=holddic[k]
    holddic[k]=tmp1.applymap(
        lambda x:None if (x==0.0) or (x!=x) or (x=="") or (x is None) else str(x)
        )
tmp2,tmp7,j=0,0,0
del greatdf,greatlis,i,k,she,shtlis,tmp1,tmp2,tmp7,j,df,dirlist,dirpath,filename,wb



#二、读取fundholdpct
dirpath=r"D:\desktop\mydatabase\fundholdpct"
dirlist=os.listdir(dirpath)

greatlis=dict()
for filename in dirlist:
    print(filename)
    #break
    if os.path.isdir(os.path.join(dirpath,filename)):
        continue
    elif  filename=="ok.xlsx":
        wb=app.books.open(os.path.join(dirpath,filename))
        wb.sheets[0].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[0].range('A1').options(pd.DataFrame,index=False,expand="table").value
        ok["重仓股占比"]=df
        continue
    wb=app.books.open(os.path.join(dirpath,filename))
    #wb[wb.sheetnames[0]].title
    greatdf=dict()
    for she in range(len(wb.sheets)):
        wb.sheets[she].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[she].range('A1').options(pd.DataFrame,index=False,expand="table").value
        greatdf[wb.sheets[she].name]=df
    greatlis[filename]=greatdf
    wb.close()

tmp1=set()
for i in greatlis.keys():
    tmp1=set(greatlis[i].keys()) | tmp1
shtlis=copy.deepcopy(tmp1)
del tmp1
holdpctdic=dict()
for sht in shtlis:
    #break
    df=pd.DataFrame(columns=["date"])
    for sce in greatlis.values():
        #break
        if sht in sce.keys():
            tmp1=copy.deepcopy(sce[sht])
            tmp1.index=pd.to_datetime(tmp1['date'])
            tmp1=tmp1.resample('3M',axis=0,closed="right",label="right").last()
            tmp1['date']=tmp1.index
            tmp1.index.name='index'
            df=pd.merge(df,tmp1,how="outer",on="date")
            tmp1=df.copy()
            i=sht
            tmp2=pd.Series(tmp1.columns).apply(lambda x:True if x[-2:-1]=="_" else False)
            if len(tmp2[tmp2])>0:
                tmp3=tmp1.loc[:,tmp1.columns[tmp2]]
                tmp4=pd.Series(tmp3.columns).apply(lambda x:x[:-2])
                tmp4=set(tmp4)
                tmp7=pd.DataFrame(index=tmp1.index)
                for j in tmp4:
                    #break
                    tmp5=tmp3.loc[:,j+"_x"]
                    tmp6=tmp3.loc[:,j+"_y"]
                    tmp5=tmp5.replace(0,np.nan)
                    tmp6=tmp6.replace(0,np.nan)
                    for k in range(len(tmp5)):
                        #break
                       if tmp5[k]==tmp5[k] or (tmp5[k]==""):
                            pass
                       else:
                            tmp5[k]=tmp6[k]
                    tmp5=pd.DataFrame(tmp5,index=tmp1.index)
                    tmp5.columns=[j]
                    tmp7=pd.concat([tmp7,tmp5],axis=1)
                tmp1=tmp1.loc[:,tmp1.columns[tmp2.apply(lambda x:not(x))]]
                del tmp2,tmp3,tmp4,tmp5,tmp6
                tmp7=pd.concat([tmp1['date'],tmp7],axis=1)
                tmp1=pd.concat([tmp7,tmp1.iloc[:,1:]],axis=1)
                df=tmp1.copy()
        else:
            pass
    df=df.sort_values(by='date',axis=0,ascending=True,inplace=False,na_position='last')
    df=df.reset_index(drop=True)
    holdpctdic[sht]=df
del sht,sce,tmp1
for k in holdpctdic.keys():
    #break
    tmp1=holdpctdic[k]
    holdpctdic[k]=tmp1.applymap(
        lambda x:None if (x==0.0) or (x!=x) or (x=="") or (x is None) else x
        )
tmp2,tmp7,j=0,0,0
del greatdf,greatlis,i,k,she,shtlis,tmp1,tmp2,tmp7,j,dirpath,dirlist,df,filename,wb

#三、读取code2indus
dirpath=r"D:\desktop\mydatabase\code2indus"
dirlist=os.listdir(dirpath)

greatdf=pd.DataFrame()
for filename in dirlist:
    print(filename)
    #break
    wb=app.books.open(os.path.join(dirpath,filename))
    #wb[wb.sheetnames[0]].title
    wb.sheets[0].used_range.last_cell.row
    wb.sheets[0].used_range.last_cell.column
    df=wb.sheets[0].range('A1').options(pd.DataFrame,index=False,expand="table").value
    df.columns=['code','indus']
    greatdf=pd.concat([greatdf,df],axis=0)
    wb.close()
code2indus=copy.deepcopy(greatdf)
del greatdf

#四、读取indusname
dirpath=r"D:\desktop\mydatabase\indusname"
dirlist=os.listdir(dirpath)

greatdf=pd.DataFrame()
for filename in dirlist:
    print(filename)
    #break
    wb=app.books.open(os.path.join(dirpath,filename))
    #wb[wb.sheetnames[0]].title
    wb.sheets[0].used_range.last_cell.row
    wb.sheets[0].used_range.last_cell.column
    df=wb.sheets[0].range('A1').options(pd.DataFrame,index=False,expand="table").value
    df.columns=['code','indusname']
    greatdf=pd.concat([greatdf,df],axis=0)
    wb.close()
codename=copy.deepcopy(greatdf)
del greatdf,df

#五、计算每个季度的基金流入红黑榜
#剔除数据不全的基金
tmp1=0
tmp2=set()
flag=0
for i in ok.values():
    flag+=1
    if flag==1:
        tmp2=set(i["fundname"])
    else:
        tmp2=set(i["fundname"]) & tmp2
fundlis=pd.Series(list(tmp2)).sort_values(ascending=True).reset_index(drop=True)
fundlis.name="fund"
del tmp1,tmp2,ok,i,flag

#算出季度红黑
induslist=list(set(code2indus['indus'])-set([None]))
ac=dict()
count=0
for i in fundlis:
    count=count+1
    print(count/len(fundlis))
    for j in range(0,10):
        #break
        holddic["第"+str(j+1)+"大重仓股代码"].loc[:,['date',i]]
        holdpctdic["第"+str(j+1)+"大重仓股占比"].loc[:,['date',i]]
        tmp1=holddic["第"+str(j+1)+"大重仓股代码"].loc[:,i]==holddic["第"+str(j+1)+"大重仓股代码"].loc[:,i]
        if len(tmp1[tmp1].index)==0:
            continue
        else:
            st=tmp1[tmp1].index[0]
        for k in range(st,len(tmp1)):
            tmp2=holddic["第"+str(j+1)+"大重仓股代码"].loc[k,i]
            tmp3=holdpctdic["第"+str(j+1)+"大重仓股占比"].loc[k,i]
            #是None的话就跳过
            if tmp2==tmp2 and not(tmp2 is None):
                pass
            else:
                continue
            
            #不是字符串的话改成字符串
            if isinstance(tmp2,str):
                pass
            else:
                tmp2="%d" % tmp2
            
            #加wind代码后缀SZSHHK
            if len(tmp2)==6 and tmp2[0]!="8" and tmp2[0]!="4":
                if tmp2[0]=="6":
                    tmp2=tmp2+".SH"
                else:
                    tmp2=tmp2+".SZ"
            elif len(tmp2)==4:
                tmp2=tmp2+".HK"
                
            tmp4=(code2indus[code2indus['code']==tmp2]['indus'].iat[0]
                  if len(code2indus[code2indus['code']==tmp2]['indus'])>0
                      else "continue"
                 )
            if tmp4=="continue":
                continue
            if k==st:
                continue
            
            datetmp=holddic["第"+str(j+1)+"大重仓股代码"]['date'][k]
            
            if datetmp in ac.keys():
                pass
            else:
                ac[datetmp]=dict()
                
            if i in ac[datetmp].keys():
                pass
            else:
                ac[datetmp][i]=dict()

            if tmp4 in ac[datetmp][i].keys():
                ac[datetmp][i][tmp4]=ac[datetmp][i][tmp4]+tmp3
            else:
                ac[datetmp][i][tmp4]=tmp3
del tmp1,tmp2,tmp3,tmp4,st,i,j,k,dirpath,dirlist,filename,datetmp,wb

datelis=pd.Series(ac.keys())
datelis.name="date"
datelis=datelis.sort_values(ascending=True).reset_index(drop=True)
tmp1=pd.MultiIndex.from_product([datelis,fundlis],names=["date","fund"])
followerpts=pd.DataFrame(None,index=tmp1,columns=["pts"])
del tmp1
for i in range(1,len(datelis.index)):
    chge=dict()
    for j in ac[datelis[i]].keys():
        #break
        if j in ac[datelis[i-1]].keys():
            pass
        else:
            continue
        tmp2=list(set(ac[datelis[i]][j].keys()) | set(ac[datelis[i-1]][j].keys()))
        for k in tmp2:
            #break
            if k in chge.keys():
                pass
            else:
                chge[k]=0.0
            if k in ac[datelis[i]][j].keys():
                if k in ac[datelis[i-1]][j].keys():
                    chge[k]=ac[datelis[i]][j][k]-ac[datelis[i-1]][j][k]
                else:
                    chge[k]=ac[datelis[i]][j][k]
            else:
                chge[k]=0-ac[datelis[i-1]][j][k]
        print(j)
    tmp1=pd.DataFrame(chge.items())
    tmp1.index=tmp1[0]
    tmp2=tmp1.rank(axis=0,method="average")[1].copy()
    for j in ac[datelis[i-1]].keys():
        #break
        tmp3=pd.Series(ac[datelis[i-1]][j].values()).sum()
        tmp4=0
        
        if j in ac[datelis[i-1]].keys():
            pass
        else:
            followerpts.loc[(datelis[i],j),"pts"]=np.nan
            continue
        
        for k in ac[datelis[i-1]][j].keys():
            #break
            ac[datelis[i-1]][j][k]
            tmp4=tmp4+tmp2.loc[k]*ac[datelis[i-1]][j][k]/tmp3
        followerpts.loc[(datelis[i],j),"pts"]=tmp4
del chge,i,j,k,tmp1,tmp2,tmp3,tmp4

followerpts.loc[(slice(None),"000006.OF"),:]
tmp1=followerpts.loc[(datelis[1],slice(None)),:]

shuchu=app.books.add()
shuchu.sheets.add()
shuchu.sheets[0].name="followerpts"
shuchu.sheets[0].range('A1').options(
    pd.DataFrame, index=True,headers=True
    ).value=followerpts

shuchu.save(r'D:\desktop\crowd\followerpts.xlsx')
app.kill()

#==================================================
#五点五、读取基金净值（此时使用index=2可以读取multiindex！！！！！）
import pandas as pd
import numpy as np
import copy
import openpyxl as op
import os
import xlwings as xw
import matplotlib.pyplot as plt

app=xw.App(visible=False,add_book=False)
app.display_alerts=False
app.screen_updating=False

wb=app.books.open(r'D:\desktop\crowd\followerpts.xlsx')
wb.sheets[0].name
wb.sheets[0].used_range.last_cell.row
wb.sheets[0].used_range.last_cell.column
followerpts=wb.sheets[0].range('A1').options(pd.DataFrame,index=2,headers=True,expand="table").value
wb.close()

#六、读取基金净值
dirpath=r"D:\desktop\mydatabase\fundnav"
dirlist=os.listdir(dirpath)

greatlis=dict()
ok=dict()
for filename in dirlist:
    print(filename)
    #break
    if os.path.isdir(os.path.join(dirpath,filename)):
        continue
    elif  filename=="ok.xlsx":
        wb=app.books.open(os.path.join(dirpath,filename))
        wb.sheets[0].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[0].range('A1').options(pd.DataFrame,index=False,expand="table").value
        ok["基金净值"]=df
        wb.close()
        continue
    wb=app.books.open(os.path.join(dirpath,filename))
    #wb[wb.sheetnames[0]].title
    greatdf=dict()
    for she in range(len(wb.sheets)):
        wb.sheets[she].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[she].range('A1').options(pd.DataFrame,index=False,expand="table").value
        greatdf["基金净值"]=df
    greatlis[filename]=greatdf
    wb.close()

tmp1=set()
for i in greatlis.keys():
    tmp1=set(greatlis[i].keys()) | tmp1
shtlis=copy.deepcopy(tmp1)
del tmp1
fundnav=dict()
for sht in shtlis:
    #break
    df=pd.DataFrame(columns=["date"])
    for sce in greatlis.values():
        #break
        print(sce.keys())
        if sht in sce.keys():
            tmp1=copy.deepcopy(sce[sht])
            tmp1.index=pd.to_datetime(tmp1['date'])
            #tmp1=tmp1.resample('3M',axis=0,closed="right",label="right").last()
            tmp1['date']=tmp1.index
            tmp1.index.name='index'
            df=pd.merge(df,tmp1,how="outer",on="date")
            tmp1=df.copy()
            i=sht
            tmp2=pd.Series(tmp1.columns).apply(lambda x:True if x[-2:-1]=="_" else False)
            if len(tmp2[tmp2])>0:
                tmp3=tmp1.loc[:,tmp1.columns[tmp2]]
                tmp4=pd.Series(tmp3.columns).apply(lambda x:x[:-2])
                tmp4=set(tmp4)
                tmp7=pd.DataFrame(index=tmp1.index)
                for j in tmp4:
                    #break
                    tmp5=tmp3.loc[:,j+"_x"]
                    tmp6=tmp3.loc[:,j+"_y"]
                    tmp5=tmp5.replace(0,np.nan)
                    tmp6=tmp6.replace(0,np.nan)
                    for k in range(len(tmp5)):
                        #break
                        if (tmp5[k]==tmp5[k]) or (tmp5[k]==""):
                            pass
                        else:
                            tmp5[k]=tmp6[k]
                    tmp5=pd.DataFrame(tmp5,index=tmp1.index)
                    tmp5.columns=[j]
                    tmp7=pd.concat([tmp7,tmp5],axis=1)
                tmp1=tmp1.loc[:,tmp1.columns[tmp2.apply(lambda x:not(x))]]
                del tmp2,tmp3,tmp4,tmp5,tmp6
                tmp7=pd.concat([tmp1['date'],tmp7],axis=1)
                tmp1=pd.concat([tmp7,tmp1.iloc[:,1:]],axis=1)
                df=tmp1.copy()
        else:
            pass
    df=df.sort_values(by='date',axis=0,ascending=True,inplace=False,na_position='last')
    df=df.reset_index(drop=True)
    fundnav[sht]=df
del sht,sce,tmp1
for k in fundnav.keys():
    #break
    tmp1=fundnav[k]
    fundnav[k]=tmp1.applymap(
        lambda x:None if (x==0.0) or (x!=x) or (x=="") or (x is None) else str(x)
        )
j=""
del greatdf,greatlis,i,k,she,shtlis,tmp1,tmp2,j,df,dirlist,dirpath,filename,wb

followerpts.loc[(slice(None),"010715.OF"),:]
followerpts.index.get_level_values(0)
followerpts.index.values

datelis=pd.Series(list(set(followerpts.index.get_level_values(level=0))))
datelis=datelis.sort_values(ascending=True).reset_index(drop=True)
datelis.name="date"
fundlis=ok["基金净值"]["fundname"]
fundlis=fundlis.sort_values(ascending=True).reset_index(drop=True)
fundlis.name="fundcode"

fundnav=fundnav["基金净值"]
fundnav.index=pd.to_datetime(fundnav["date"])
del fundnav["date"]
fundnav=fundnav.resample('D',axis=0,closed="right",label="right").bfill()

mypf=pd.DataFrame(0.0,index=datelis,columns=["value","share","holding"])
st=20
for dt in range(st,len(datelis)):
    if dt==st:
        mypf.loc[datelis[dt],"value"]=1
        mypf.loc[datelis[dt],"share"]=0
        mypf.loc[datelis[dt],"holding"]="allcash"
        continue
    tmp1=followerpts.loc[(datelis[dt],slice(None)),slice(None)]
    tmp1=tmp1.droplevel(level=0,axis=0).copy().sort_values(by="pts",ascending=True)
    tmp1.dropna(axis=0,how="all",inplace=True)
    for i in range(len(tmp1.index)):
        tmp3=tmp1.index[i]
        if tmp3 in set(fundlis):
            pass
        else:
            continue
        if tmp3 in fundnav.columns:
            tmp4=fundnav.loc[datelis[dt-1],tmp3]
        else:
            continue
        if (tmp4==tmp4) and not(tmp4 is None):
            tmp2=tmp3
            break
    tmp3=mypf.loc[datelis[dt-1],"holding"]
    if (i==len(tmp1.index)-1):
        if tmp3!="allcash":
            mypf.loc[datelis[dt],"value"]=(
                float(fundnav.loc[datelis[dt],tmp3]) / 
                float(fundnav.loc[datelis[dt-1],tmp3])
                )*mypf.loc[datelis[dt-1],"value"]
            mypf.loc[datelis[dt],"share"]=mypf.loc[datelis[dt-1],"share"]
            mypf.loc[datelis[dt],"holding"]=tmp2
            continue
        else:
            mypf.loc[datelis[dt],"value"]=mypf.loc[datelis[dt-1],"value"]
            mypf.loc[datelis[dt],"share"]=0
            mypf.loc[datelis[dt],"holding"]="allcash"
            continue
    if tmp3=="allcash":
        mypf.loc[datelis[dt],"value"]=mypf.loc[datelis[dt-1],"value"]
        mypf.loc[datelis[dt],"share"]=(
            mypf.loc[datelis[dt],"value"]/float(fundnav.loc[datelis[dt-1],tmp2])
            )
        mypf.loc[datelis[dt],"holding"]=tmp2
    else:
        mypf.loc[datelis[dt],"value"]=(
            float(fundnav.loc[datelis[dt],tmp3]) / 
            float(fundnav.loc[datelis[dt-1],tmp3])
            )*mypf.loc[datelis[dt-1],"value"]
        mypf.loc[datelis[dt],"share"]=(
            mypf.loc[datelis[dt],"value"]/float(fundnav.loc[datelis[dt-1],tmp2])
            )
        mypf.loc[datelis[dt],"holding"]=tmp2
        
mypf=mypf.iloc[st:,]
del dt,followerpts,i,ok,st,tmp1,tmp2,tmp3,tmp4

comparefd=fundnav.loc[mypf.index,:].copy()
comparefd=comparefd["885001.WI"]
comparefd=comparefd.apply(lambda x: float(x)/float(comparefd.iat[0]))

plt.rcParams["font.sans-serif"]=["SimHei"] #设置字体
plt.rcParams["axes.unicode_minus"]=False #该语句解决图像中的“-”负号的乱码问题
#fig,ax=plt.subplots(2,1)
#ax[0][0]
fig,ax=plt.subplots(1,1)
ax.plot(mypf.index,mypf['value'])
ax.plot(mypf.index,comparefd)
ax.legend(labels = ('MyPortfolio','偏股混合指数'), loc = 'lower right') # legend placed at lower right
ax.set_title("净值曲线")
ax.set_xlabel('Date')
ax.set_ylabel('NAV')

plt.show()

mypf.to_excel(r'D:\desktop\crowd\myportfolio.xlsx',header=True,index=True)

#app.quit()
app.kill()














